# app/services/receipt_service.py
from sqlalchemy.orm import Session
from sqlalchemy import text
import openpyxl
import shutil
import os
import zipfile
from openpyxl.styles import Alignment
from datetime import datetime

from ..core.config import settings

class ReceiptService:
    """수령증 생성 서비스 (기존 로직 100% 보존)"""
    
    def __init__(self, db: Session):
        self.db = db
    
    def generate_receipts_for_user(self, user_id: int) -> str:
        """사용자별 수령증 생성 (기존 get_matched_name_and_payback 로직)"""
        output_dir = f"{settings.OUTPUT_DIR}/user_{user_id}"
        
        # 기존 디렉토리가 있다면 삭제
        if os.path.exists(output_dir):
            shutil.rmtree(output_dir)
        
        os.makedirs(output_dir, exist_ok=True)
        
        # 사용자의 면세점 타입을 동적으로 감지
        duty_free_type = self._detect_duty_free_type(user_id)
        print(f"사용자 {user_id}의 면세점 타입: {duty_free_type}")
        
        if duty_free_type == "lotte":
            # 롯데 면세점 데이터 조회 (동적 테이블)
            sql1 = text("""
            SELECT e.name, e."PayBack"
            FROM lotte_excel_data e
            JOIN receipt_match_log m
            ON e."receiptNumber" = m.receipt_number
            WHERE m.is_matched = TRUE AND m.user_id = :user_id;
            """)
        else:
            # 신라 면세점 데이터 조회 (동적 테이블)
            sql1 = text("""
            SELECT e.name, e."PayBack"
            FROM shilla_excel_data e
            JOIN receipt_match_log m
            ON e."receiptNumber"::text = m.receipt_number
            WHERE m.is_matched = TRUE AND m.user_id = :user_id;
            """)

        sql2 = text("""
        SELECT p.passport_number, p.birthday, p.name
        FROM passports p
        WHERE p.name = :name AND p.user_id = :user_id;
        """)

        try:
            results = self.db.execute(sql1, {"user_id": user_id}).fetchall()
            print(f"매칭된 엑셀 데이터 조회 결과: {len(results)}건")
        except Exception as e:
            print(f"엑셀 데이터 조회 오류: {e}")
            results = []
            
        printed_people = set()
        generated_count = 0

        for name, payback in results:
            try:
                passport_result = self.db.execute(
                    sql2, {"name": name, "user_id": user_id}
                ).fetchone()

                if passport_result:
                    passport_number, birthday, passport_name = passport_result
                    person = (passport_name, payback, passport_number, birthday)

                    if person in printed_people:
                        continue
                    printed_people.add(person)

                    output_path = os.path.join(output_dir, f"{passport_name}_수령증.xlsx")
                    shutil.copy(settings.RECEIPT_TEMPLATE_PATH, output_path)

                    wb = openpyxl.load_workbook(output_path)
                    ws = wb.active
                    alignment = Alignment(horizontal="center", vertical="center")
                    
                    ws["D7"] = passport_name
                    ws["D7"].alignment = alignment
                    ws["D8"] = passport_number
                    ws["D8"].alignment = alignment
                    ws["D9"] = birthday
                    ws["D9"].alignment = alignment
                    ws["D10"] = payback
                    ws["D10"].alignment = alignment
                    
                    today = datetime.today()
                    formatted_date = f"{today.year}년    {today.month:02}월    {today.day:02}일"
                    ws["B16"] = formatted_date
                    ws["B16"].alignment = alignment
                    
                    wb.save(output_path)
                    generated_count += 1
                    
                    print(f"수령증 생성 완료: {passport_name}")
                else:
                    print(f"여권 정보를 찾을 수 없음: {name}")
            except Exception as e:
                print(f"개별 수령증 생성 오류: {e}")
                continue
        
        if generated_count == 0:
            raise Exception("생성된 수령증이 없습니다. 매칭된 데이터를 확인해주세요.")
        
        # ZIP 파일로 압축
        zip_path = f"{settings.OUTPUT_DIR}/user_{user_id}_receipts.zip"
        with zipfile.ZipFile(zip_path, 'w') as zipf:
            for root, dirs, files in os.walk(output_dir):
                for file in files:
                    if file.endswith('.xlsx'):
                        file_path = os.path.join(root, file)
                        arcname = file
                        zipf.write(file_path, arcname)
                        print(f"ZIP에 추가: {file}")
        
        # 임시 파일들 정리
        shutil.rmtree(output_dir)
        
        print(f"총 {generated_count}개의 수령증 생성 완료")
        print(f"수령증 ZIP 파일 생성 완료: {zip_path}")
        
        return zip_path
    
    def _detect_duty_free_type(self, user_id: int) -> str:
        """사용자의 현재 데이터를 기반으로 면세점 타입을 감지 (기존 로직 보존)"""
        try:
            # 신라 데이터 개수 확인
            shilla_count_sql = text("SELECT COUNT(*) FROM shilla_receipts WHERE user_id = :user_id")
            shilla_count = self.db.execute(shilla_count_sql, {"user_id": user_id}).scalar() or 0
            
            # 롯데 데이터 개수 확인
            lotte_count_sql = text("SELECT COUNT(*) FROM receipts WHERE user_id = :user_id")
            lotte_count = self.db.execute(lotte_count_sql, {"user_id": user_id}).scalar() or 0
            
            print(f"데이터 감지: 신라={shilla_count}, 롯데={lotte_count}")
            
            # 더 많은 데이터가 있는 쪽을 선택
            if shilla_count >= lotte_count:
                return "shilla"
            else:
                return "lotte"
                
        except Exception as e:
            print(f"면세점 타입 감지 오류: {e}")
            # 오류 시 기본값 반환
            return "lotte"